#!/usr/bin/env python3
"""
extract_schedule.py

Lê a planilha "SEMANA PADRÃO" (.xlsx) e gera um JSON estruturado de eventos
para as categorias:
  - HAM (palestra + prática, todos os grupos)
  - CI Prática (Clínica Integrada - CI Prática, todos os grupos/locais)
  - PIEPE (todos os grupos)
  - IESC / Comunidades - Prática (todos os grupos) + IESC - Palestra (comum)
  - Clínica Integrada - CI MARC + CI Palestra (comum a todos os grupos)

Qualquer outra categoria/disciplina da planilha (ex. ÁREA VERDE, MOMENTO NED)
é IGNORADA propositalmente, conforme escopo definido pelo usuário.

Uso:
    python3 extract_schedule.py [caminho_para_xlsx] [caminho_json_saida]

Padrão:
    entrada:  ../data/schedule.xlsx  (relativo a este script)
    saida:    ../output/events.json
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

# --------------------------------------------------------------------------
# Configuração
# --------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR.parent / "data" / "schedule.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "output" / "events.json"

# Colunas (coluna de horário, coluna de grupo/preceptor) por dia da semana,
# na ordem em que aparecem na planilha "SEMANA PADRÃO".
DAY_COLUMNS = [
    ("SEGUNDA", "A", "B"),
    ("TERCA", "C", "D"),
    ("QUARTA", "E", "F"),
    ("QUINTA", "G", "H"),
    ("SEXTA", "I", "J"),
]

DAY_LABELS = {
    "SEGUNDA": "Segunda-feira",
    "TERCA": "Terça-feira",
    "QUARTA": "Quarta-feira",
    "QUINTA": "Quinta-feira",
    "SEXTA": "Sexta-feira",
}

TIME_RE = re.compile(r"(\d{1,2}):(\d{1,2})\s*-\s*(\d{1,2}):(\d{1,2})")


def fmt_hhmm(h, m):
    return f"{int(h):02d}:{int(m):02d}"


def strip_accents(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def norm(s):
    """Normaliza texto para comparação: maiúsculas, sem acento, espaços colapsados."""
    if s is None:
        return ""
    s = strip_accents(str(s)).upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s


# --------------------------------------------------------------------------
# Classificação de categorias (SÓ estas são extraídas — todo o resto é ignorado)
# --------------------------------------------------------------------------

def classify_category(type_label_norm, type_label_raw=""):
    """Recebe o rótulo do "tipo de atividade" (já normalizado) e devolve
    (categoria, subtipo) ou None se a categoria não faz parte do escopo pedido.
    type_label_raw (com acentos/maiúsculas originais) é usado só para exibição."""

    if type_label_norm.startswith("HAM - PALESTRA"):
        return ("HAM", "Palestra")
    if type_label_norm.startswith("HAM - PRATICA") or type_label_norm.startswith("HAM - PRÁTICA".upper()):
        return ("HAM", "Prática")
    if type_label_norm.startswith("HAM"):
        # fallback genérico para variações de grafia do HAM
        return ("HAM", "Outro")

    if type_label_norm.startswith("CI PRATICA"):
        # extrai o local depois do primeiro " - ", preservando a grafia original
        raw_parts = type_label_raw.split(" - ", 1)
        location = raw_parts[1].strip() if len(raw_parts) > 1 else ""
        return ("CI_PRATICA", location)

    if type_label_norm == "PIEPE":
        return ("PIEPE", None)

    if type_label_norm.startswith("COMUNIDADES - PRATICA"):
        return ("IESC_COMUNIDADES", "Prática")
    if type_label_norm.startswith("IESC - PALESTRA"):
        return ("IESC_COMUNIDADES", "Palestra")

    if "CLINICA INTEGRADA" in type_label_norm and "MARC" in type_label_norm:
        return ("CI_MARC_PALESTRA", "CI MARC")
    if "CLINICA INTEGRADA" in type_label_norm and "PALESTRA" in type_label_norm:
        return ("CI_MARC_PALESTRA", "CI Palestra")

    return None  # fora do escopo pedido (ex: AREA VERDE, MOMENTO NED, etc.)


CATEGORY_DISPLAY = {
    "HAM": "HAM",
    "CI_PRATICA": "Clínica Integrada – CI Prática",
    "PIEPE": "PIEPE",
    "IESC_COMUNIDADES": "IESC / Comunidades",
    "CI_MARC_PALESTRA": "Clínica Integrada – CI MARC / Palestra",
}

# --------------------------------------------------------------------------
# Parsing de grupo + preceptor a partir do texto livre da célula
# --------------------------------------------------------------------------

def parse_group_and_preceptor(text):
    text_norm = re.sub(r"\s+", " ", text).strip()
    text_norm = text_norm.strip(" -")

    # "TODOS - Prof. Fulano"  -> comum a todos os grupos
    m = re.match(r"(?i)^todos\s*-?\s*(.*)$", text_norm)
    if m:
        return {"group": "TODOS", "preceptor": m.group(1).strip(), "raw": text_norm}

    # "GRUPO 1 - Fulano" / "Grupo 01 - Fulano"
    m = re.match(r"(?i)^grupo\s*0*(\d+)\s*-\s*(.+)$", text_norm)
    if m:
        return {"group": m.group(1), "preceptor": m.group(2).strip(), "raw": text_norm}

    # "Fulano - 5/6" / "Fulano -5/6" / "Fulano - Grupos 7/8"
    m = re.match(r"(?i)^(.+?)\s*-\s*(?:grupos?\s*)?(\d+)\s*/\s*(\d+)\s*$", text_norm)
    if m:
        n1, n2 = int(m.group(2)), int(m.group(3))
        group = f"{min(n1, n2)}/{max(n1, n2)}"
        return {"group": group, "preceptor": m.group(1).strip(), "raw": text_norm}

    # "Fulano 9/10" (sem hífen antes do grupo)
    m = re.match(r"(?i)^(.+?)\s+(\d+)\s*/\s*(\d+)\s*$", text_norm)
    if m:
        n1, n2 = int(m.group(2)), int(m.group(3))
        group = f"{min(n1, n2)}/{max(n1, n2)}"
        return {"group": group, "preceptor": m.group(1).strip(), "raw": text_norm}

    # Sem grupo identificável (ex: preceptores de CI MARC/Palestra) -> comum
    return {"group": None, "preceptor": text_norm, "raw": text_norm}


# --------------------------------------------------------------------------
# Extração principal
# --------------------------------------------------------------------------

def extract_events(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    events = []
    warnings = []
    time_typos = []
    event_id = 0

    for day_key, col1, col2 in DAY_COLUMNS:
        current_type_raw = None
        current_type_norm = None

        for row in range(1, ws.max_row + 1):
            v1 = ws[f"{col1}{row}"].value
            v2 = ws[f"{col2}{row}"].value

            if v1 is None and v2 is None:
                continue

            v1s = str(v1).strip() if v1 is not None else ""

            time_match = TIME_RE.match(v1s)

            if time_match:
                # Linha de DETALHE (horário + grupo/preceptor)
                if current_type_norm is None:
                    continue  # detalhe sem cabeçalho antes (não deveria ocorrer)

                cat = classify_category(current_type_norm, current_type_raw)
                if cat is None:
                    continue  # fora do escopo pedido -> ignora

                category, subtype = cat

                start, end = fmt_hhmm(time_match.group(1), time_match.group(2)), fmt_hhmm(time_match.group(3), time_match.group(4))
                # só é typo de verdade se algum componente tiver 1 dígito só
                # (ex: "18:0" em vez de "18:00"); diferenças de espaço/traço
                # no separador são só formatação, não erro de dado.
                if any(len(g) == 1 for g in time_match.groups()):
                    time_typos.append(
                        f'{DAY_LABELS[day_key]} - {current_type_raw} (cél. {col1}{row}): '
                        f'planilha tem "{time_match.group(0)}", interpretado como "{start} - {end}" — confirme se está certo'
                    )

                # o texto do grupo/preceptor pode estar na coluna 2, ou
                # (caso de HAM-PALESTRA) embutido na mesma célula da coluna 1
                remainder_col1 = v1s[time_match.end():].strip()
                if v2:
                    detail_text = str(v2).strip()
                elif remainder_col1:
                    detail_text = remainder_col1
                else:
                    detail_text = ""

                parsed = parse_group_and_preceptor(detail_text) if detail_text else {
                    "group": None, "preceptor": None, "raw": ""
                }

                event_id += 1
                events.append({
                    "id": f"ev{event_id:03d}",
                    "day": day_key,
                    "day_label": DAY_LABELS[day_key],
                    "category": category,
                    "category_label": CATEGORY_DISPLAY[category],
                    "subtype": subtype,
                    "type_label_raw": current_type_raw,
                    "start": start,
                    "end": end,
                    "group": parsed["group"],
                    "preceptor": parsed["preceptor"],
                    "raw_detail": parsed["raw"],
                    "cell": f"{col1}{row}/{col2}{row}",
                })
            else:
                # Linha de TIPO/CABEÇALHO (define a categoria das próximas
                # linhas de detalhe, até o próximo cabeçalho)
                current_type_raw = v1s
                current_type_norm = norm(v1s)

    # ------------------------------------------------------------------
    # Checagem de ambiguidades: mesmo dia + mesma categoria + mesmo horário
    # com preceptores/grupos diferentes, ou horários sobrepostos.
    # ------------------------------------------------------------------
    def to_minutes(hhmm):
        h, m = hhmm.split(":")
        return int(h) * 60 + int(m)

    by_day = {}
    for ev in events:
        by_day.setdefault(ev["day"], []).append(ev)

    critical = []   # mesmo preceptor em dois lugares ao mesmo tempo -> conflito real
    info = []       # horários idênticos/sobrepostos mas com preceptores diferentes -> normalmente
                    # são grupos rodando em paralelo (esperado na CI Prática), mas listamos
                    # para o usuário revisar caso ache que não deveria ser assim.

    for day, evs in by_day.items():
        for i in range(len(evs)):
            for j in range(i + 1, len(evs)):
                a, b = evs[i], evs[j]
                if a["category"] != b["category"] or a["subtype"] != b["subtype"]:
                    continue

                a1, a2 = to_minutes(a["start"]), to_minutes(a["end"])
                b1, b2 = to_minutes(b["start"]), to_minutes(b["end"])
                identical_time = a["start"] == b["start"] and a["end"] == b["end"]
                overlapping = a1 < b2 and b1 < a2
                if not identical_time and not overlapping:
                    continue

                same_preceptor = (
                    a["preceptor"] and b["preceptor"]
                    and a["preceptor"].strip().lower() == b["preceptor"].strip().lower()
                )
                same_location = a["type_label_raw"] == b["type_label_raw"]

                msg = (
                    f'{DAY_LABELS[day]} - "{a["category_label"]}" ({a["type_label_raw"]}): '
                    f'{a["start"]}-{a["end"]} [grupo {a["group"]}, {a["preceptor"]}] '
                    f'{"==" if identical_time else "sobrepoe"} '
                    f'{b["start"]}-{b["end"]} [grupo {b["group"]}, {b["preceptor"]}]'
                )

                if same_preceptor:
                    critical.append(msg + "  <- MESMO PRECEPTOR em dois horarios que se cruzam (conflito real)")
                elif identical_time and same_location and not same_preceptor:
                    info.append(msg + "  (mesmo local/horario, preceptores diferentes - provavel grupos paralelos)")
                else:
                    info.append(msg)

    return events, critical, info, time_typos


def build_group_options(events):
    """Constrói, por categoria filtrável, a lista ordenada de grupos disponíveis."""
    options = {}
    for cat in ("HAM", "CI_PRATICA", "PIEPE", "IESC_COMUNIDADES"):
        groups = set()
        for ev in events:
            if ev["category"] == cat and ev["group"] and ev["group"] != "TODOS":
                groups.add(ev["group"])

        def sort_key(g):
            first = g.split("/")[0]
            try:
                return (0, int(first))
            except ValueError:
                return (1, first)

        options[cat] = sorted(groups, key=sort_key)
    return options


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not xlsx_path.exists():
        print(f"ERRO: arquivo não encontrado: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    events, critical, info, time_typos = extract_events(xlsx_path)
    group_options = build_group_options(events)

    out_data = {
        "source_file": xlsx_path.name,
        "generated_events": len(events),
        "default_filters": {
            "HAM": "1",
            "CI_PRATICA": "5/6",
            "PIEPE": "1",
            "IESC_COMUNIDADES": "2",
        },
        "group_options": group_options,
        "category_labels": CATEGORY_DISPLAY,
        "events": events,
        "warnings": {
            "conflitos_criticos": critical,
            "notas_informativas": info,
            "horarios_fora_do_padrao": time_typos,
        },
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out_data, f, ensure_ascii=False, indent=2)

    # Também grava uma versão .js (window.SCHEDULE_DATA = {...};) para que o
    # index.html carregue os dados via <script src>, o que funciona mesmo
    # abrindo o HTML localmente (file://), sem os bloqueios de CORS que
    # fetch()/XHR têm para arquivos locais.
    js_path = out_path.with_suffix(".js")
    with open(js_path, "w", encoding="utf-8") as f:
        f.write("// Gerado automaticamente por scripts/extract_schedule.py — não edite manualmente.\n")
        f.write("window.SCHEDULE_DATA = ")
        json.dump(out_data, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    print(f"OK: {len(events)} eventos extraidos -> {out_path}")
    print(f"OK: dados tambem gravados em -> {js_path}")
    print("\nGrupos disponiveis por categoria filtravel:")
    for cat, groups in group_options.items():
        print(f"  {CATEGORY_DISPLAY[cat]}: {groups}")

    if critical:
        print(f"\n[CRITICO] {len(critical)} conflito(s) real(is) (mesmo preceptor, horarios cruzados):")
        for w in critical:
            print(f"  - {w}")
    else:
        print("\nNenhum conflito critico (mesmo preceptor em horarios cruzados) detectado.")

    if time_typos:
        print(f"\n[ATENCAO] {len(time_typos)} horario(s) fora do padrao HH:MM na planilha (interpretados, mas CONFIRME):")
        for w in time_typos:
            print(f"  - {w}")

    if info:
        print(f"\n[INFO] {len(info)} coincidencia(s) de horario (provavelmente grupos paralelos, revisar se quiser):")
        for w in info:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
